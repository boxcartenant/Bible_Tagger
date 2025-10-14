import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog, Misc, StringVar, OptionMenu
import bibledb_lib as bdblib
from tkinter.font import Font
from openpyxl import Workbook
import os
from collections import defaultdict
import numpy as np
import numpy as np
import matplotlib.pyplot as plt

def combineVRefs(vref1, vref2):
    #get the verse reference and store in self.current_data['ref']
    #start book and chapter
    bookparts = vref1.split(":")[0].split(" ")
    bA = ""
    i = 0
    while i < len(bookparts)-1:
       bA += bookparts[i] + " "
       i+= 1
    bA = bA.strip()
    cA = vref1.split(":")[0].split(" ")[i]
    bookparts = vref2.split(":")[0].split(" ")
    #end book and chapter
    bB = ""
    i = 0
    while i < len(bookparts)-1:
       bB += bookparts[i] + " "
       i+= 1
    bB = bB.strip()
    cB = vref2.split(":")[0].split(" ")[i]
    #start and end verses
    try:
        vA = vref1.split(":")[1]
        vB = vref2.split(":")[1]
    except Exception as e:
        print("Error in combineVRefs(): ",e)
        print("vref1 =", vref1)
        print("vref2 =", vref2)
        return vref1
    if bA != bB: #different book
        if bdblib.book_proper_names.index(bA) < bdblib.book_proper_names.index(bB):
            return(bA+" "+cA + ":" + vA + " - " + bB + " " +cB + ":" + vB)
        else:
            return(bB+" "+cB + ":" + vB + " - " + bA + " " +cA + ":" + vA)
    elif cA != cB: #same book, different chapter
        if int(cA) < int(cB):
            return(bA+" "+cA + ":" + vA + "-" + cB + ":" + vB)
        else:
            return(bA+" "+cB + ":" + vB + "-" + cA + ":" + vA)
    elif vA != vB: #same book, same chapter, different verse
        if int(vA) < int(vB):
            return(bA+" "+cA + ":" + vA + "-" + vB)
        else:
            return(bA+" "+cB + ":" + vB + "-" + vA)
    else: #same book, same chapter, same verse
        return vref1

def bargraph_size(test, testmin, testmax, barmin=0, barmax=100):
    test_adjusted = test - testmin
    testmax_adjusted = testmax - testmin
    
    # Handle case where all values are the same (avoid division by zero)
    if testmax_adjusted == 0:
        result = barmin  # or could use (barmin + barmax) / 2 for middle value
    else:
        result = (test_adjusted * (barmax - barmin) / testmax_adjusted) + barmin
    
    if result < 3:
        result = 3
    return result

def color_gradient(length, min_length, max_length, type_selection):
    if type_selection == "redblue":
        # Red to Blue pastel gradient
        r = 255-int(bargraph_size(length, min_length, max_length, 0, 255))
        g = int(bargraph_size(length, min_length, max_length, 50, 180))
        b = int(bargraph_size(length, min_length, max_length, 0, 255))
    elif type_selection == "purpleyellow":
        # Purple to Yellow pastel gradient
        r = int(bargraph_size(length, min_length, max_length, 200, 255))
        g = 255-int(bargraph_size(length, min_length, max_length, 0, 255))
        b = int(bargraph_size(length, min_length, max_length, 0, 255))
    else:
        return 'azure'
    # Return color in hexadecimal format
    return f'#{r:02x}{g:02x}{b:02x}'


class TagInputDialog(simpledialog.Dialog):
    #inherits simpledialog to make a tag input dialog with a dropdown suggestion list
    def __init__(self, parent, dbdata, topselection = False, get_tags_like=bdblib.get_tags_like, thistitle="Add Tag", bookinputdialog=False):
        #if topselection is false, entering a new tagname will return the new tagname
        #if topselection is true, entering a new tagname will return the top item in the list box
        #get_tags_like is a callback to whatever function you intend to use to find similar tags. (maybe unnecessary for this to be an argument)
        self.topselection = topselection
        self.get_tags_like = get_tags_like
        self.selected_tag = None
        self.dbdata = dbdata
        self.bookinputdialog = bookinputdialog
        super().__init__(parent, title=thistitle)

    def body(self, master):
        # Entry for typing tag
        self.entry = tk.Entry(master)
        self.entry.grid(row=0, column=0, padx=10, pady=10)
        # Listbox to display tag suggestions
        self.listbox = tk.Listbox(master, height=5)
        self.listbox.grid(row=1, column=0, padx=10, pady=10)
        self.listbox.grid_remove()  # Initially hide the listbox
        # Bind entry events for filtering suggestions
        self.entry.bind("<KeyRelease>", self.update_suggestions)
        self.entry.bind("<Down>", self.focus_listbox)  # Bind down arrow key
        self.listbox.bind("<Double-1>", self.on_select)
        self.listbox.bind('<Button-1>', self.quick_select)
        self.listbox.bind("<Return>", self.on_select)  # Bind enter key in listbox
        return self.entry  # Focus on entry widget

    def update_suggestions(self, event):
        # Get the current input from the entry widget
        partial_tag = self.entry.get()

        # Clear the listbox
        self.listbox.delete(0, tk.END)
        if partial_tag.strip():
            # Get suggestions from the database
            matching_tags = self.get_tags_like(self.dbdata, partial_tag)
            if matching_tags:
                # Insert suggestions into the listbox
                for tag in matching_tags:
                    self.listbox.insert(tk.END, tag)

                # Show the listbox if there are matching tags
                self.listbox.grid()
            else:
                # Hide the listbox if no matches are found
                self.listbox.grid_remove()
        else:
            # Hide the listbox if the input is empty
            self.listbox.grid_remove()

    def focus_listbox(self, event):
        # If there are items in the listbox, select the first one and focus the listbox
        if self.listbox.size() > 0:
            self.listbox.selection_set(0)  # Select the first item
            self.listbox.activate(0)       # Make the first item active
            self.listbox.focus_set()       # Move focus to the listbox

    def quick_select(self, event):
        #print("quick select")
        index = self.listbox.nearest(event.y)
        if index >= 0:
            if self.bookinputdialog:
                selected_text = self.listbox.get(index)
            else:
                selected_text = self.listbox.get(index)[0]
            self.entry.delete(0,tk.END)
            self.entry.insert(0,selected_text)
    
    def on_select(self, event):
        #print("on select")
        # Get the selected tag from the listbox
        selected = self.listbox.curselection()
        if self.bookinputdialog and selected:
            self.selected_tag = self.listbox.get(selected[0])
            self.ok()
        elif selected:
            self.selected_tag = self.listbox.get(selected[0])[0]
            self.ok()  # Close the dialog

    def apply(self):
        #print("apply")
        selected = None
        # Get the final tag value (from entry or listbox)
        if self.topselection and self.listbox.size() > 0:
            # Check if there are no highlighted items
            has_selected_items = any(self.listbox.selection_includes(i) for i in range(self.listbox.size()))
            
            # Check if there are no items identical to the string from self.entry.get()
            search_string = self.entry.get()
            #print(search_string, [self.listbox.get(i) for i in range(self.listbox.size())])
            #if self.bookinputdialog:
            #    matching_index = next((i for i in range(self.listbox.size()) if self.listbox.get(i) == search_string), None)
            #else:
            matching_index = next((i for i in range(self.listbox.size()) if self.listbox.get(i)[0] == search_string), None)
    
            # If no highlighted items and no matching items, select the first item
            if not has_selected_items:
                if matching_index is not None:
                    #print("found!")
                    # If an identical item is found, select it
                    self.listbox.selection_set(matching_index)
                    self.listbox.activate(matching_index)  # Make the matching item active
                else:
                    # If no matches, select the first item
                    self.listbox.selection_set(0)  # Select the first item
                    self.listbox.activate(0)       # Make the first item active
                self.listbox.focus_set()       # Move focus to the listbox
                selected = self.listbox.curselection()
        if selected and self.bookinputdialog:
            self.selected_tag = self.listbox.get(selected[0])
        elif selected:
            self.selected_tag = self.listbox.get(selected[0])[0]
        elif not self.selected_tag:
            self.selected_tag = self.entry.get()   

######################Secondary Window is the window and the bargraph column. The filter view is instantiated here.

class DBExplorer:
    def __init__(self, master, callback, dbdata = None):
        self.master = master
        self.callback = callback #used for clicking a tag and setting it as the current tag in the main window
        self.dbdata = dbdata #info about the currend db that's open
        self.top_window = None
        
        #Handy variables here...
        self.canvasFont = Font(size = 10)
        self.italicFont = Font(size = 10, slant = 'italic')
        self.boldFont = Font(size = 10, weight = 'bold')
        self.sortmode_text = "Sorting Alphabetically"
        self.sortmode = "alphabet" #toggles between "alphabet" and "usage"
        self.colormode_text = "Toggle Colors"
        self.colormode = "plain" #toggles: "plain", "redblue", "purpleyellow"
        self.export_tags_this_time = False
        self.export_tags_path = "./"
        self.all_tags_list = [] # this variable contains the tags sorted by usage. It is used by right_hand_frame to export tag/note/verses 

    def show(self, callback = None, dbdata = None):
        self.callback = callback #used for clicking a tag and setting it as the current tag in the main window
        self.dbdata = dbdata #info about the currend db that's open
        if self.top_window and self.top_window.winfo_exists():
            self.top_window.lift()
            return
        # Set up the secondary window
        self.top_window = tk.Toplevel(self.master)
        self.top_window.title("DB Explorer")
        self.top_window.iconbitmap("./bibletaggericon.ico")
        self.top_window.geometry("800x500")
        self.reload_id = None
        self.reload_id2 = None

        def on_close():
            self.top_window.destroy()
            self.top_window = None
        self.top_window.protocol("WM_DELETE_WINDOW", on_close)

        self.populate()
        


    def populate(self):
        #I am accustomed to dealing with paned windows
        self.this_window = ttk.PanedWindow(self.top_window, orient="horizontal")
        self.this_window.pack(fill="both", expand=True)

        #Left-hand pane
        self.myFrame = ttk.Frame(self.this_window)
        self.myFrame.grid(row=0, column=0, sticky="nsew")
        self.myFrame.grid_rowconfigure(0, weight=1)
        self.myFrame.grid_columnconfigure(0, weight=1)
        
        #add the canvas to the frame
        self.canvas = tk.Canvas(self.myFrame)
        self.canvas.grid(row=0, column=0, sticky="nsew")

        #create vertical and horizontal scrollbars
        v_scrollbar = tk.Scrollbar(self.myFrame, orient="vertical", command=self.canvas.yview)
        h_scrollbar = tk.Scrollbar(self.myFrame, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Configure scrollbar position on the grid
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")

        #add the frame to the window
        self.this_window.add(self.myFrame)

        #add another frame to the right-hand side of the window. We're using two columns now, baby!
        self.verse_sorting_panel = VerseSortingPanel(self.master, self.top_window, self.this_window, self.callback, self.dbdata, self)
        self.verse_sorting_panel.grid(row=0, column=1, sticky="nsew")
        self.this_window.add(self.verse_sorting_panel)

        #configure the scroll region to make the canvas scrollable
        canvas_width = self.canvas.winfo_reqwidth()
        canvas_height = self.canvas.winfo_reqheight()
        self.canvas.config(scrollregion=(0, 0, canvas_width, canvas_height))

        #whenever the window resizes, I want the canvas scroll area to refresh
        self.canvas.bind("<Configure>", self.on_resize)

        #finish rendering the window before showing attributes, so that the dimensions will be accurate
        self.top_window.update_idletasks()
        
        #sash drag for resizing the panels
        self.this_window.bind("<B1-Motion>", self.on_sash_drag)
        self.this_window.bind("<Configure>", self.window_resize)

        #go ahead and display contents
        self.top_window.after(50, self.display_attributes)

        self.active_panel = None
        self.top_window.bind("<MouseWheel>", self.scroll_active_panel)
        self.canvas.bind("<Enter>", self.set_active_panel)
        self.canvas.bind("<Leave>", self.clear_active_panel)
        self.verse_sorting_panel.canvas.bind("<Enter>", self.set_active_panel)
        self.verse_sorting_panel.canvas.bind("<Leave>", self.clear_active_panel)

    def set_active_panel(self, event):
        # Set the active panel to the widget under the mouse
        self.active_panel = event.widget
        #print("active_panel",self.active_panel)

    def clear_active_panel(self, event):
        # Clear the active panel when the mouse leaves
        self.active_panel = None

    def scroll_active_panel(self, event):
        # Only scroll the active panel
        #print("scrolling active_panel",self.active_panel)
        if self.active_panel:
            first, last = self.active_panel.yview()  # Get scrollbar position
            if event.delta > 0 and first <= 0:  # Trying to scroll up but already at the top
                return  # Do nothing
            if event.delta < 0 and last >= 1:  # Trying to scroll down but already at the bottom
                return
            self.active_panel.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def on_resize(self, event):
        #this is for instances where the bargraph canvas is resized.

        #reloading is resource intensive for this window, so we don't want to do for every new window size.
        # Cancel any existing timer
        if self.reload_id is not None:
            self.canvas.after_cancel(self.reload_id)

        # Set a new timer to reload attributes after 300 ms of inactivity
        self.reload_id = self.canvas.after(300, self.display_attributes)

    def window_resize(self, event):
        # Resizing the window doesn't move the sash, but only results in resizing the right-hand frame.
        self.verse_sorting_panel.display_attributes(canvas_width = self.this_window.winfo_width() - self.this_window.sashpos(0))
    
    def on_sash_drag(self,event):
        # If the moved sash is near the treeview sash (left), update the tree view
        if self.this_window.sashpos(0) - 10 < event.x < self.this_window.sashpos(0) + 10:
            new_sash_position = event.x
            self.myFrame.columnconfigure(0, weight=1)
            self.myFrame.configure(width=new_sash_position)
            #update the canvas
            self.verse_sorting_panel.display_attributes(canvas_width = self.this_window.winfo_width() - new_sash_position)
            self.on_resize(event)
            #self.verse_sorting_panel.display_attributes(canvas_width = self.this_window.winfo_width() - new_sash_position)
    
    def on_tag_click(self, event, item_id):
        self.callback(item_id)
        #here is where we will callback to the main window to show tag data

    #I know string compares are heavy, but this is a light application overall at the moment.
    def on_sortmode_click(self, event):
        #this is for a button to toggle between...
        # a) sorting tags in alphabetical order (every tag gets a line, synonyms are shown multiple times)
        # b) sorting tags according to usage (synonym groups get just one line)
        if self.sortmode == "alphabet": #toggles between "alphabet" and "usage"
            self.sortmode = "usage"
            self.sortmode_text = "Sorting by usage"
        else:
            self.sortmode = "alphabet"
            self.sortmode_text = "Sorting Alphabetically"
        self.display_attributes()

    def on_colormode_click(self, event):
        #this is for a button to toggle between...
        # a) showing tags on a color gradient for usage (maybe a few options)
        # b) showing all tags in the same color
        if self.colormode == "plain":#toggles: "plain", "redblue", "purpleyellow"
            self.colormode = "redblue"
        elif self.colormode == "redblue":
            self.colormode = "purpleyellow"
        else:
            self.colormode = "plain"
        self.display_attributes()

    def on_exporttags_click(self, event):
        self.export_tags_this_time = True
        self.export_tags_path = filedialog.asksaveasfilename(defaultextension=".xls", filetypes=[("Excel", "*.xls"), ("All files", "*.*")])
        if self.export_tags_path:
            self.export_tags_this_time = True
        self.display_attributes()

    def show_loading_overlay(self, text="Loading…"):
        """Create or raise a loading overlay on the canvas."""
        cx = self.canvas.winfo_width() // 2
        cy = self.canvas.winfo_height() // 2
        if not hasattr(self, 'loading_text_id') or self.loading_text_id is None:
            self.loading_text_id = self.canvas.create_text(
                cx, cy, text=text, anchor=tk.CENTER,
                font=self.boldFont, fill="gray", tags="loading_overlay"
            )
        else:
            self.canvas.itemconfigure(self.loading_text_id, text=text)
            self.canvas.coords(self.loading_text_id, cx, cy)
            self.canvas.tag_raise(self.loading_text_id)
        self.canvas.update_idletasks()  # force it to draw immediately

    def hide_loading_overlay(self):
        if hasattr(self, 'loading_text_id') and self.loading_text_id:
            self.canvas.delete(self.loading_text_id)
            self.loading_text_id = None


    def display_attributes(self, dbstuff = None):
        self.show_loading_overlay("Loading…")
        #shows all the tags
        x_offset = 5
        y_offset = 3
        textelbowroom = 10
        textlinegap = 2
        textlineheight = Font.metrics(self.canvasFont)["linespace"]
        boldlineheight = Font.metrics(self.boldFont)["linespace"]
        italiclineheight = Font.metrics(self.italicFont)["linespace"]
        right_x_offset = 30
        panelWidth = self.this_window.sashpos(0) - right_x_offset

        #Clear the canvas
        self.canvas.delete("all")
        
        if dbstuff != None:
            self.dbdata = dbstuff
        else:
            dbstuff = self.dbdata

        if self.dbdata is None:
            self.canvas.create_text(10, 30, text="Select a verse to load a DB.", fill="green", font = self.canvasFont)
        else:
            #get all the tags
            checklist = [b['tag'] for b in bdblib.get_tag_list(self.dbdata)] #all the tags
            
            syngroups = []   #each indice is a list of synonymous tags
            synonymlist = [] #a group of synonymous tags to be added to syngroups
            checkedlist = [] #tags we've already dealt with. So we don't deal with them again.
            synonyms = []    #a working list of tags we're in the middle of finding synonyms for
            
            for tag in checklist:
                if tag not in checkedlist:
                    checkedlist.append(tag)
                    synonymlist.append(tag)
                    synonyms = [b['tag'] for b in bdblib.get_db_stuff(self.dbdata,"tag","tag",tag)]
                    while len(synonyms) > 0:
                        synonym = synonyms.pop()
                        if synonym not in checkedlist:
                            checkedlist.append(synonym)
                            synonymlist.append(synonym)
                            synonyms += [b['tag'] for b in bdblib.get_db_stuff(self.dbdata,"tag","tag",synonym)]
                    
                  #  #OLD Recursive way of getting syngroups
                  #  def recursivesynonyms(syno, thisgroup):
                  #      synonyms = [b['tag'] for b in bdblib.get_db_stuff(self.dbdata,"tag","tag",syno)]
                  #      for tag in synonyms:
                  #          if tag not in thisgroup:
                  #              thisgroup.append(tag)
                  #              thisgroup = recursivesynonyms(tag, thisgroup)
                  #      return thisgroup
                  #  
                  #  synonymlist = recursivesynonyms(tag, [tag])
                  #  
                  #  for s in synonymlist:
                  #      if s != tag:
                  #          checkedlist.append(s)
                  #      
                    syngroups.append({"tags":synonymlist})
                    synonymlist = []
                

            syngroups_lo = 9223372036854775807
            syngroups_hi = 0
            #count the unique verses for the synonym group
            #these low and high values are used to scale bargraph lengths and colors
            verses = []
            for i in range(len(syngroups)):
                for tag in syngroups[i]["tags"]:
                    checkverses = bdblib.get_db_stuff(self.dbdata, "verse", "tag", tag)
                    for verse in checkverses:
                        if verse not in verses:
                            verses.append(verse)
                syngroups[i]["verses"] = len(verses)
                if syngroups[i]["verses"] < syngroups_lo:
                    syngroups_lo = syngroups[i]["verses"]
                if syngroups[i]["verses"] > syngroups_hi:
                    syngroups_hi = syngroups[i]["verses"]
                verses = []
                                

            tags_lo = 9223372036854775807
            tags_hi = 0
            tags = [{"tag":b} for b in checklist]
            #count the verses for each tag
            for i in range(len(tags)):
                tags[i]["verses"] = len(bdblib.get_db_stuff(self.dbdata, "verse", "tag", tags[i]["tag"]))
                if tags[i]["verses"] < tags_lo:
                    tags_lo = tags[i]["verses"]
                if tags[i]["verses"] > tags_hi:
                    tags_hi = tags[i]["verses"]
            
            #show some statistics at the top
            #  qty of tags
            showtext = "total tag count: " +str(len(tags))
            self.canvas.create_text(x_offset+textelbowroom, y_offset+textelbowroom, text=showtext, anchor=tk.NW, font=self.canvasFont)
            y_offset += textlineheight + textlinegap
            #  qty of synonym groups
            showtext = "tag count after grouping synonyms: " +str(len(syngroups))
            self.canvas.create_text(x_offset+textelbowroom, y_offset+textelbowroom, text=showtext, anchor=tk.NW, font=self.canvasFont)
            y_offset += textelbowroom + textlineheight + textlinegap
            
            #show sorting buttons
            buttonText = self.sortmode_text #"Toggle Sorting"
            sort_button_width = self.canvasFont.measure(buttonText) + 2*textelbowroom
            self.canvas.create_rectangle(x_offset, y_offset, x_offset+sort_button_width,y_offset + textlineheight + 2*textelbowroom, fill='snow', tags='on_sortmode_click')
            self.canvas.create_text(x_offset+textelbowroom, y_offset+textelbowroom, text=buttonText, anchor=tk.NW, font=self.canvasFont, tags='on_sortmode_click')
            self.canvas.tag_bind('on_sortmode_click', '<Button-1>', self.on_sortmode_click)

            

            buttonText = self.colormode_text #"Toggle Coloring"
            color_button_width = self.canvasFont.measure(buttonText) + 2*textelbowroom
            color_x_offset = x_offset
            if x_offset + sort_button_width + textelbowroom*2 + color_button_width < panelWidth:
                color_x_offset += sort_button_width + textelbowroom*2
            else:
                y_offset += 10 + textlineheight + 2*textelbowroom
            self.canvas.create_rectangle(color_x_offset, y_offset, color_x_offset+color_button_width,y_offset + textlineheight + 2*textelbowroom, fill='snow', tags='on_colormode_click')
            self.canvas.create_text(color_x_offset+textelbowroom, y_offset+textelbowroom, text=buttonText, anchor=tk.NW, font=self.canvasFont, tags='on_colormode_click')
            self.canvas.tag_bind('on_colormode_click', '<Button-1>', self.on_colormode_click)

            buttonText = "Export List"
            export_button_width = self.canvasFont.measure(buttonText) + 2*textelbowroom
            export_x_offset = color_x_offset
            if export_x_offset + textelbowroom*2 + color_button_width + export_button_width < panelWidth:
                export_x_offset += color_button_width + textelbowroom*2
            else:
                export_x_offset = x_offset
                y_offset += 10 + textlineheight + 2*textelbowroom
            self.canvas.create_rectangle(export_x_offset, y_offset, export_x_offset+export_button_width,y_offset + textlineheight + 2*textelbowroom, fill='snow', tags='on_exporttags_click')
            self.canvas.create_text(export_x_offset+textelbowroom, y_offset+textelbowroom, text=buttonText, anchor=tk.NW, font=self.canvasFont, tags='on_exporttags_click')
            self.canvas.tag_bind('on_exporttags_click', '<Button-1>', self.on_exporttags_click)

            y_offset += 10 + textlineheight + 2*textelbowroom
            
            #now we have...
            # tags -- ["tagname","tagname1","tagname2",...]
            # tagverses -- [versecount, versecount1, versecount2,...]
            # syngroups -- [{"tags":["tagname","tagname1","tagname2",...],"verses":integer_verse_count},...]
            
            # tags_hi -- the qty of references to the most used tag
            # tags_lo -- the qty of references to the least used tag
            # syngroups_hi -- the qty of references to the most used syngroup
            # syngroups_lo -- the qty of references to the least used syngroup
            #... the ranges for tags might be inaccurate in view of syngroups where another synonym has the references

            #self.sortmode_text = "sorting"#the button text
            #self.sortmode = "alphabet" #toggles between "alphabet" and "usage"
            #self.colormode_text = "color" #the button text
            #self.colormode = "plain" #toggles: "plain", "redblue", "purpleyellow"

            #prepare the workbook
            workbook = None
            sheet = None
            if self.export_tags_this_time:
                workbook = Workbook()
                sheet = workbook.active
                wbheader = ["Verse Count", "Tags..."]
                sheet.append(wbheader)
                
            if self.sortmode == "alphabet":
                sorted_tags = sorted(tags, key=lambda x: x["tag"])
                sorted_syngroups = [(syngroups.index(lst), lst["tags"].index(dct["tag"])) for dct in sorted_tags for lst in syngroups if dct["tag"] in lst["tags"]]
            elif self.sortmode == "usage":
                #We are only using syngroups for display, so no need to sort tags here.
                #sorted_tags = sorted(tags, key=lambda x: x["verses"], reverse = True) #reverse to put the most-used tags on top
                sorted_syngroups = sorted(syngroups, key=lambda x: x["verses"], reverse = True)
            tagnum = 0
            self.all_tags_list = [s["tags"] for s in sorted(syngroups, key=lambda x: x["verses"], reverse = True)]# this variable is used by right_hand_frame to export tag/note/verses 
            for s in sorted_syngroups:
                #print(s)
                if self.sortmode == "alphabet":
                    #in this case, s is a list of tuples, representing the index of the tag in syngroups.
                    tags = [syngroups[s[0]]["tags"][s[1]]]
                    verses = syngroups[s[0]]["verses"]
                elif self.sortmode == "usage":
                    #in this case, s is a dict, like {'tags': ['beard', 'hair', 'beards'], 'verses': 4}
                    tags = s["tags"]
                    verses = s["verses"]
                if self.export_tags_this_time:
                    sheet.append([verses]+tags)

                barcolor = color_gradient(verses, syngroups_lo, syngroups_hi, self.colormode)
                barlength = bargraph_size(verses, syngroups_lo, syngroups_hi, x_offset, panelWidth)
                
                
                #use the width of the window to determine how many times this tag list will wrap
                #   use that width to determine the needed height for this bar-graph rectangle

                linesize = textelbowroom
                barheight = textlineheight + textlinegap*2
                rendertags = [] #this will store, in rows, all the tags to be displayed on this bar
                tagline = []
                #first calculate the total height of this block of tags, for the bar graph bar height
                for tag in tags:
                    tagwidth = self.canvasFont.measure(tag)
                    if tagwidth + linesize + textelbowroom < panelWidth:
                        tagline.append(tag)
                        linesize += tagwidth + textelbowroom
                    else:
                        rendertags.append(tagline)
                        tagline = [tag]
                        barheight += textlineheight + textlinegap
                        linesize = tagwidth+textelbowroom
                if tagline != []:
                    rendertags.append(tagline)

                #render the bar, and then iterate through the tags and display them
                tag_binder = "GOTO_tag_"+str(tagnum) #the bar binds to the first tag in the group
                self.canvas.create_rectangle(x_offset, y_offset, x_offset+barlength, y_offset+barheight, fill=barcolor, tags=tag_binder)
                y_offset += textlinegap
                for tagline in rendertags:
                    tagx = textelbowroom
                    for i, tag in enumerate(tagline):
                        tag_binder = "GOTO_tag_"+str(tagnum)
                        self.canvas.create_text(x_offset+tagx, y_offset, text=tag, anchor=tk.NW, font=self.canvasFont, tags=tag_binder)
                        self.canvas.tag_bind(tag_binder, '<Button-1>', lambda event, item_id = tag : self.on_tag_click(event,item_id))
                        tagnum += 1
                        tagwidth = self.canvasFont.measure(tag)
                        tagx += tagwidth + textelbowroom
                        if i < len(tagline)-1:
                            self.canvas.create_line(tagx, y_offset-(textlinegap/2), tagx, y_offset+textlineheight+(textlinegap/2)+2, fill='black', width=2)
                            #the +2 in the second y coord above was just an asthetic tweak, and can be removed if you don't like it.
                            tagx += 2
                    y_offset += textlineheight + textlinegap
                y_offset += 10
        if self.export_tags_this_time:
            self.export_tags_this_time = False
            workbook.save(self.export_tags_path)
            print("Exported tags to: " + str(self.export_tags_path))
            self.export_tags_path = None
        self.hide_loading_overlay()
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        

###########################################Right Hand Frame is the verse sorting area. Probably not a useful name for it, now that I think about it.

class VerseSortingPanel(ttk.Frame):
    def __init__(self, master, parent_window, parent, callback = None, dbdata = None, left_frame = None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.parent_window = parent_window
        self.canvasFont = Font(size = 10)
        self.italicFont = Font(size = 10, slant = 'italic')
        self.boldFont = Font(size = 10, weight = 'bold')
        self.left_frame = left_frame
        
        self.callback = callback
        self.dbdata = dbdata
        
        # Configure grid inside this frame
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Create a canvas inside the right-hand frame
        self.canvas = tk.Canvas(self)
        self.canvas.grid(row=0, column=0, sticky="nsew")

        self.canvas.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))

        # Create vertical and horizontal scrollbars
        self.v_scrollbar = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.h_scrollbar = tk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)

        # Configure canvas to use scrollbars
        self.canvas.configure(yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set)

        # Add scrollbars to the canvas frame
        self.v_scrollbar.grid(row=0, column=1, sticky="ns")
        self.h_scrollbar.grid(row=1, column=0, sticky="ew")

        # Handy variables
        self.tags_list = []
        self.union = True #Show all of the verses for the selected tags
        self.intersection = False #Show only the verses shared by all selected tags
        self.symmetric_difference = False #Show only the verses that are unique to only one of the selected tags
        self.verse_xref_list = []
        self.selected_books = []
        self.comments_only = False #only show verses with comments
        self.shown_verses = [] #in case you want to do something with the verses that are currently visible
        

        # Configure the scroll region to make the canvas scrollable
        self.canvas_width = self.canvas.winfo_reqwidth()
        self.canvas_height = self.canvas.winfo_reqheight()
        self.canvas.config(scrollregion=(0, 0, self.canvas_width, self.canvas_height))
        
        # the rest of the stuff
        self.display_attributes()
        
        #self.canvas.create_text(10, 10, anchor="nw", text="Right-hand canvas content")

    def get_books_like(self, dbdata, partial_book):
        result = [book for book in bdblib.book_proper_names if partial_book.lower().strip() in book.lower().strip()]
        return result
    
    def select_book(self, event):
        selected_book = TagInputDialog(self.master, self.dbdata, topselection = True, get_tags_like = self.get_books_like, thistitle = "Select Books", bookinputdialog = True).selected_tag
        if (selected_book is not None) and (selected_book != "") and (selected_book in bdblib.book_proper_names) and (selected_book not in self.selected_books):
            self.selected_books.append(selected_book)
            self.display_attributes()

    def delete_book(self, event, bookname):
        if bookname in self.selected_books:
            self.selected_books.remove(bookname)
            self.display_attributes()
        else:
            print("Tried removing a book that wasn't listed from the filter pane. Oh no!")

    def select_tag(self, event):
        #replace tagSelect() with the actual code for tag selection, to get the popup.
        selected_tag = TagInputDialog(self.master, self.dbdata, topselection = True).selected_tag
        if (selected_tag is not None) and (selected_tag != "") and bdblib.tag_exists(self.dbdata, selected_tag):
            selected_tag_synonyms = [b['tag'] for b in bdblib.get_db_stuff(self.dbdata,"tag","tag",selected_tag)]
            
            if selected_tag and (selected_tag not in self.tags_list) and all(item not in self.tags_list for item in selected_tag_synonyms):
                self.tags_list.append(selected_tag)
                self.display_attributes()
            
    def delete_tag(self, event, tagname):
        if tagname in self.tags_list:
            self.tags_list.remove(tagname)
            self.display_attributes()
        else:
            print("Tried removing a tag that wasn't listed from the filter pane. Oh no!")

    def on_tag_click(self, event, item_id):
        #this callback goes to the main window to show tag data
        self.callback(item_id)

    def on_verse_click(self, event, item_id):
        #this callback goes to the main window to show tag data
        self.callback(item_id, "verse")

    def toggle_union(self, event):
        if self.union:
            self.union = False
            self.intersection = True
        elif self.intersection:
            self.intersection = False
            self.symmetric_difference = True
        elif self.symmetric_difference:
            self.symmetric_difference = False
            self.union = True
        self.display_attributes()

    def toggle_comments_only(self, wasted_memory):
        self.comments_only = not self.comments_only
        self.display_attributes()

    def copy_verse_lists(self, event):
        #puts the currently visible verse list in the clipboard
        result = ""
        for verse in self.shown_verses:
            result += verse
        root = tk._default_root
        if root:
            root.clipboard_clear()
            root.clipboard_append(result)
            root.update()
        else:
            print("ERROR in bdblib_Manager.copy_verse_list(). No access to root.")

    def export_tags_and_synonyms(self, event):

        def get_verses_for_taglist(tags):
            result = []
            for synonym_group in tags:
                verses = []
                notes = []
                for tag in synonym_group:
                    this_note = bdblib.get_db_stuff(self.dbdata, "note", "tag", tag)
                    if this_note:
                        this_note = this_note[0]['note']
                        notes.append(this_note)
                    checkverses = bdblib.get_db_stuff(self.dbdata, "verse", "tag", tag)
                    for verse in checkverses:
                        if verse not in verses:
                            #print(verse)
                            verses.append(verse)
                #I don't know if these functions will work, but they're basically what's being done down below in display_verses
                verses.sort(key=lambda r: (r['start_book'], r['start_chapter'], r['start_verse']))#sort by start book first, then chapter, then verse, so all the verses appear in order
                verses = [bdblib.normalize_vref(q) for q in verses]
                result.append({"tags":synonym_group,"notes":notes, "verses":verses})
            return result
        
        #self.dbdata
        tagslist = self.left_frame.all_tags_list
        exportdata = get_verses_for_taglist(tagslist)
        contents = ""
        for item in exportdata:
            contents += '\n  //////TAGS://////\n'
            for tag in item['tags']:
                contents += tag + ", "
            contents += '\n  //////NOTES://////\n'
            for note in item['notes']:
                contents += note + '\n\n'
            contents += '\n  //////VERSES://////\n'
            for verse in item['verses']:
                contents += verse + ', '
            contents += "\n\n+=+=+=+=+=+=+=+=+=+=+=+=+=+=+\n\n"
                
        output_file = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text Files", "*.txt"), ("All files", "*.*")])
        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(contents)
                    
    def export_verse_notes(self, event): ##################################################
        #self.dbdata
        verses = bdblib.get_all_verses_with_notes(self.dbdata)
        contents = ""
        for verse in verses:
            contents += verse['verse'] + '\n'
            contents += verse['note']
            contents += "\n\n+=+=+=+=+=+=+=+=+=+=+=+=+=+=+\n\n"
        output_file = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text Files", "*.txt"), ("All files", "*.*")])
        if output_file:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(contents)
        
    def export_tag_networks(self, event):
        out_dir = filedialog.askdirectory(title="Select folder for tag network output")
        if not out_dir:
            return
        messagebox.showinfo("Information", "This will take a long time. Press OK and then wait for another popup.")
        self.parent_window.lift()
        subfolder_name = "tag networks"
        folder = os.path.join(out_dir, subfolder_name)
        os.makedirs(folder, exist_ok=True)
        tagslist = self.left_frame.all_tags_list
        def get_verses_for_taglist(tags):
            result = []
            for synonym_group in tags:
                verses = []
                for tag in synonym_group:
                    checkverses = bdblib.get_db_stuff(self.dbdata, "verse", "tag", tag)
                    for verse in checkverses:
                        if verse not in verses:
                            verses.append(verse)
                verses.sort(key=lambda r: (r['start_book'], r['start_chapter'], r['start_verse']))
                verse_refs = [bdblib.normalize_vref(q) for q in verses]
                result.append({"tags":synonym_group, "verses":verse_refs})
            return result
        exportdata = get_verses_for_taglist(tagslist)
        for item in exportdata:
            primary_tag = item['tags'][0].replace("/", "_").replace("\\", "_")
            filename = os.path.join(folder, primary_tag + ".txt")
            cooccurring_tags = set()
            for vref in item['verses']:
                tags_on_verse = [t['tag'] for t in bdblib.get_db_stuff(self.dbdata, "tag", "verse", vref)]
                cooccurring_tags.update(tags_on_verse)
            contents = ", ".join(sorted(cooccurring_tags)) + "\n"
            contents += "----------\n"
            for vref in item['verses']:
                tags_on_verse = [t['tag'] for t in bdblib.get_db_stuff(self.dbdata, "tag", "verse", vref)]
                contents += vref + "\n"
                contents += ", ".join(tags_on_verse) + "\n\n"
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(contents)
        messagebox.showinfo("Information", f"Exported files to folder, '{folder}'!")
        self.parent_window.lift()

    def export_subtopic_breakdowns(self, event):
        out_dir = filedialog.askdirectory(title="Select folder for subtopic breakdowns")
        if not out_dir:
            return
        messagebox.showinfo("Information", "This will take a long time. Press OK and then wait for another popup.")
        self.parent_window.lift()
        subfolder_name = "subtopic breakdowns"
        folder = os.path.join(out_dir, subfolder_name)
        os.makedirs(folder, exist_ok=True)
        tagslist = self.left_frame.all_tags_list
        for synonym_group in tagslist:
            verses = []
            for tag in synonym_group:
                checkverses = bdblib.get_db_stuff(self.dbdata, "verse", "tag", tag)
                for verse in checkverses:
                    if verse not in verses:
                        verses.append(verse)
            verses.sort(key=lambda r: (r['start_book'], r['start_chapter'], r['start_verse']))
            verse_refs = [bdblib.normalize_vref(q) for q in verses]
            verse_objs = verses
            verse_dict = dict(zip(verse_refs, verse_objs))
            verse_tags = {}
            for vref in verse_refs:
                tags = [t['tag'] for t in bdblib.get_db_stuff(self.dbdata, "tag", "verse", vref)]
                verse_tags[vref] = set(tags)
            co_tags = set()
            for tags in verse_tags.values():
                co_tags.update(tags)
            checklist = list(co_tags)
            sub_syngroups = []
            synonymlist = []
            checkedlist = []
            synonyms = []
            for tag in checklist:
                if tag not in checkedlist:
                    checkedlist.append(tag)
                    synonymlist.append(tag)
                    synonyms = [b['tag'] for b in bdblib.get_db_stuff(self.dbdata,"tag","tag",tag)]
                    while len(synonyms) > 0:
                        synonym = synonyms.pop()
                        if synonym not in checkedlist:
                            checkedlist.append(synonym)
                            synonymlist.append(synonym)
                            synonyms += [b['tag'] for b in bdblib.get_db_stuff(self.dbdata,"tag","tag",synonym)]
                    sub_syngroups.append(synonymlist)
                    synonymlist = []
            t_set = set(synonym_group)
            sub_syngroups = [sg for sg in sub_syngroups if set(sg) != t_set]
            sub_data = []
            for sub_sg in sub_syngroups:
                s_set = set(sub_sg)
                sub_verses = [v for v, vtags in verse_tags.items() if s_set & vtags]
                sub_data.append((len(sub_verses), sub_sg, sub_verses))
            sub_data.sort(key=lambda x: x[0], reverse=True)
            for _, _, sv in sub_data:
                sv.sort(key=lambda vr: (verse_dict[vr]['start_book'], verse_dict[vr]['start_chapter'], verse_dict[vr]['start_verse'], verse_dict[vr]['end_book'], verse_dict[vr]['end_chapter'], verse_dict[vr]['end_verse']))
            primary_tag = synonym_group[0].replace("/", "_").replace("\\", "_")
            filename = os.path.join(folder, primary_tag + ".txt")
            contents = ""
            for count, sg, sverses in sub_data:
                if count > 0:
                    contents += f"///// SUBTOPIC ({count} verses): " + ", ".join(sg) + "\n"
                    contents += ", ".join(sverses) + "\n\n"
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(contents)

            # Generate heatmap for overlaps (limit to top 20 subtopics)
            if len(sub_data) > 0:
                sub_data = sub_data[:]  # Limit for visualization
                sub_sets = [set(sverses) for _, _, sverses in sub_data]
                n = len(sub_sets)
                if n > 1:  # Need at least 2 for meaningful overlaps
                    overlap = np.zeros((n, n))
                    for i in range(n):
                        for j in range(n):
                            inter = len(sub_sets[i] & sub_sets[j])
                            union = len(sub_sets[i] | sub_sets[j])
                            jacc = inter / union if union else 0
                            overlap[i, j] = jacc
                    fig, ax = plt.subplots(figsize=(10, 8))
                    im = ax.imshow(overlap, cmap='viridis')
                    ax.set_xticks(range(n))
                    ax.set_yticks(range(n))
                    labels = [sg[0] for _, sg, _ in sub_data]  # Use first tag as label
                    ax.set_xticklabels(labels, rotation=45, ha='right')
                    ax.set_yticklabels(labels)
                    plt.colorbar(im, label='Jaccard Similarity')
                    plt.title(f"Subtopic Overlap Heatmap for {primary_tag}")
                    plt.tight_layout()
                    plt.savefig(os.path.join(folder, primary_tag + "_overlap_heatmap.png"))
                    plt.close()
        messagebox.showinfo("Information", f"Exported files to folder, '{folder}'!")
        self.parent_window.lift()

    def analyze_subtopic_tags(self, event, tags=None, verses=None):
        """
        Analyze what *synonym groups* occur within a given list of verses or tags.
        Clicking a bar or its label shows verses AND copies them to clipboard.
        """

        # ---------------
        # 1. Build the verse set
        # ---------------
        verse_refs = set()

        if verses is not None and len(verses) > 0:
            for vtuple in verses:
                vref = bdblib.normalize_vref({
                    "ID": None,
                    "start_book": vtuple[0],
                    "start_chapter": vtuple[1],
                    "start_verse": vtuple[2],
                    "end_book": vtuple[3],
                    "end_chapter": vtuple[4],
                    "end_verse": vtuple[5]
                })
                verse_refs.add(vref)
        elif tags is not None and len(tags) > 0:
            # union of verses from these tags + synonyms
            for t in tags:
                stack = [t]
                seen_tags = set()
                while stack:
                    tg = stack.pop()
                    if tg in seen_tags:
                        continue
                    seen_tags.add(tg)
                    for row in bdblib.get_db_stuff(self.dbdata, "verse", "tag", tg):
                        vref = bdblib.normalize_vref(row)
                        verse_refs.add(vref)
                    # synonyms
                    syns = [b['tag'] for b in bdblib.get_db_stuff(self.dbdata, "tag", "tag", tg)]
                    for s in syns:
                        if s not in seen_tags:
                            stack.append(s)
        else:
            messagebox.showinfo("No Data", "No verses or tags to analyze.")
            self.parent_window.lift()
            return

        if not verse_refs:
            messagebox.showinfo("No Data", "No verses found for these tags.")
            self.parent_window.lift()
            return

        # ---------------
        # 2. Build synonym groups among all tags found in these verses
        # ---------------
        # First collect all tags from these verses
        all_tags_in_verses = set()
        verse_to_tags = defaultdict(set)
        for vref in verse_refs:
            rows = bdblib.get_db_stuff(self.dbdata, "tag", "verse", vref)
            for r in rows:
                t = r['tag']
                verse_to_tags[vref].add(t)
                all_tags_in_verses.add(t)

        # Build synonym groups: connected components of tag↔tag synonyms restricted to our tags
        visited = set()
        syn_groups = []
        for t in all_tags_in_verses:
            if t in visited:
                continue
            stack = [t]
            comp = set()
            while stack:
                tg = stack.pop()
                if tg in visited:
                    continue
                visited.add(tg)
                comp.add(tg)
                syns = [b['tag'] for b in bdblib.get_db_stuff(self.dbdata, "tag", "tag", tg)]
                for s in syns:
                    if s in all_tags_in_verses and s not in visited:
                        stack.append(s)
            syn_groups.append(comp)

        # Remove the input tags themselves from groups if you don’t want to count them
        #input_tag_set = set(tags) if tags else set()
        #filtered_groups = []
        #for g in syn_groups:
        #    if g & input_tag_set and len(g - input_tag_set) == 0:
        #        # group is only the input tags themselves
        #        continue
        #    filtered_groups.append(g)
        #syn_groups = filtered_groups

        if not syn_groups:
            messagebox.showinfo("No Data", "No other tag groups found in these verses.")
            self.parent_window.lift()
            return

        # For each synonym group, find verses where ANY tag in the group occurs
        group_to_verses = []
        for g in syn_groups:
            vs = {vref for vref, tagsset in verse_to_tags.items() if tagsset & g}
            group_to_verses.append((g, vs))

        # Sort groups by count of verses
        group_to_verses.sort(key=lambda x: len(x[1]), reverse=True)

        # Build display labels
        group_labels = []
        for g, vs in group_to_verses:
            first_tag = sorted(g)[0]
            label = f"{first_tag} (+{len(g)-1} synonyms)" if len(g) > 1 else first_tag
            group_labels.append(label)

        # ---------------
        # 3. Build UI with clickable bars & labels
        # ---------------
        vis_window = tk.Toplevel(self.master)
        vis_window.title("Tag-Synonym Groups in Selected Verses")
        vis_window.geometry("900x600")

        canvas_frame = ttk.Frame(vis_window)
        canvas_frame.pack(fill="both", expand=True)
        canvas = tk.Canvas(canvas_frame)
        canvas.grid(row=0, column=0, sticky="nsew")
        v_scrollbar = tk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = tk.Scrollbar(canvas_frame, orient="horizontal", command=canvas.xview)
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        canvas_font = Font(size=10)
        textlineheight = canvas_font.metrics()["linespace"]
        x_offset = 10
        y_offset = 10

        canvas.create_text(x_offset, y_offset, text="Tag-Synonym Groups in Selected Verses", anchor=tk.NW, font=canvas_font)
        y_offset += textlineheight + 12

        max_val = len(group_to_verses[0][1])
        bar_area_width = 600
        bar_height = 18
        label_col_width = 300

        # ---------------
        # 4. Click handler to show and copy verses
        # ---------------
        def show_and_copy_verses(verses, label_text):
            # Put verses on clipboard
            result = ", ".join(verses)
            root = tk._default_root
            if root:
                root.clipboard_clear()
                root.clipboard_append(result)
                root.update()
            # Show message
            if len(verses) > 200:
                sample = ", ".join(verses[:200])
                messagebox.showinfo(f"Verses for {label_text}",
                                    f"{len(verses)} verses (first 200 shown):\n{sample}\n(All verses copied to clipboard.)")
                vis_window.lift()
            else:
                messagebox.showinfo(f"Verses for {label_text}", ", ".join(verses))
                vis_window.lift()

        for i, ((group, vs), label) in enumerate(zip(group_to_verses, group_labels)):
            val = len(vs)
            lab_y = y_offset + i * (bar_height + 6)

            # Label text
            ttag = f"lab_{i}"
            canvas.create_text(x_offset, lab_y, text=label, anchor=tk.NW, font=canvas_font, tags=ttag)

            # Bar rectangle
            bar_x = x_offset + label_col_width
            bar_len = (val / max_val) * bar_area_width if max_val > 0 else 0
            btag = f"bar_{i}"
            canvas.create_rectangle(bar_x, lab_y, bar_x + bar_len, lab_y + bar_height,
                                   fill='skyblue', outline='black', tags=btag)
            canvas.create_text(bar_x + bar_len + 6, lab_y + bar_height / 2,
                              text=str(val), anchor=tk.W, font=canvas_font)

            # Bind both label and bar to the click handler
            verses_sorted = sorted(vs)
            canvas.tag_bind(ttag, "<Button-1>", lambda e, v=verses_sorted, l=label: show_and_copy_verses(v, l))
            canvas.tag_bind(btag, "<Button-1>", lambda e, v=verses_sorted, l=label: show_and_copy_verses(v, l))

        total_height = y_offset + len(group_to_verses) * (bar_height + 6) + 50
        total_width = x_offset + label_col_width + bar_area_width + 200
        canvas.configure(scrollregion=(0, 0, total_width, total_height))

        messagebox.showinfo("Done", "Click a group name or bar to view its verses.\n(All verses also copied to clipboard.)")
        self.parent_window.lift()
        vis_window.lift()



    def export_single_topics(self, event):
        out_dir = filedialog.askdirectory(title="Select folder for tag single topic output")
        if not out_dir:
            return
        messagebox.showinfo("Information", "This will take a long time. Press OK and then wait for another popup.")
        self.parent_window.lift()
        def get_verses_for_taglist(tags):
            result = []
            for synonym_group in tags:
                verses = []
                notes = []
                for tag in synonym_group:
                    this_note = bdblib.get_db_stuff(self.dbdata, "note", "tag", tag)
                    if this_note:
                        this_note = this_note[0]['note']
                        notes.append(this_note)
                    checkverses = bdblib.get_db_stuff(self.dbdata, "verse", "tag", tag)
                    for verse in checkverses:
                        if verse not in verses:
                            verses.append(verse)
                verses.sort(key=lambda r: (r['start_book'], r['start_chapter'], r['start_verse']))
                verses = [bdblib.normalize_vref(q) for q in verses]
                result.append({"tags":synonym_group,"notes":notes, "verses":verses})
            return result
        
        subfolder_name = "single topics"
        folder = os.path.join(out_dir, subfolder_name)
        os.makedirs(folder, exist_ok=True)
        tagslist = self.left_frame.all_tags_list
        exportdata = get_verses_for_taglist(tagslist)
        for item in exportdata:
            primary_tag = item['tags'][0].replace("/", "_").replace("\\", "_")
            filename = os.path.join(folder, primary_tag + ".txt")
            contents = ""
            contents += '\n  //////TAGS://////\n'
            contents += ', '.join(item['tags']) + '\n'
            contents += '  //////NOTES://////\n'
            contents += '\n\n'.join(item['notes']) + '\n\n'
            contents += '  //////VERSES://////\n'
            contents += ', '.join(item['verses']) + '\n\n'
            contents += "+=+=+=+=+=+=+=+=+=+=+=+=+=+=+\n\n"
            contents += '  //////VERSE NOTES://////\n'
            for verse in item['verses']:
                vnotes = bdblib.get_db_stuff(self.dbdata, "note", "verse", verse)
                if vnotes:
                    contents += verse + '\n'
                    contents += vnotes[0]['note'] + '\n\n'
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(contents)
        messagebox.showinfo("Information", f"Exported files to folder, '{folder}'!")
        self.parent_window.lift()

    def export_pairwise_overlaps(self, event):
        """
        Compute pairwise overlaps between all tags in the DB and export:
          - CSV with counts of overlaps between each pair of tags
          - GEXF network file (for Gephi)
        """
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
        import os, time, csv
        from collections import defaultdict
        import numpy as np
        import networkx as nx

        # ----------------------------
        # UI: progress window
        # ----------------------------
        progress_window = tk.Toplevel(self.master)
        progress_window.title("Computing Pairwise Overlaps")
        progress_window.geometry("350x110")
        progress_label = tk.Label(progress_window, text="Preparing...")
        progress_label.pack(pady=8)
        progress_bar = ttk.Progressbar(progress_window, orient="horizontal", length=300, mode="determinate")
        progress_bar.pack(pady=6)
        progress_window.update_idletasks()

        # ----------------------------
        # Step 1: Load all tags and verses once
        # ----------------------------
        progress_label.config(text="Loading all tags and verses...")
        progress_bar['value'] = 5
        progress_window.update_idletasks()

        # Build tag→verse mapping
        tag_to_verses = defaultdict(set)
        all_tags = [r['tag'] for r in bdblib.get_tag_list(self.dbdata)]
        all_tags = sorted(set(all_tags))
        for t in all_tags:
            verses = bdblib.get_db_stuff(self.dbdata, "verse", "tag", t)
            for v in verses:
                vref = bdblib.normalize_vref(v)
                tag_to_verses[t].add(vref)

        # remove empty tags
        all_tags = [t for t in all_tags if tag_to_verses[t]]

        n_tags = len(all_tags)
        if n_tags < 2:
            messagebox.showinfo("No Data", "Not enough tags with verses to compute overlaps.")
            progress_window.destroy()
            return

        # Choose output folder
        out_dir = filedialog.askdirectory(title="Select folder for overlap output")
        if not out_dir:
            progress_window.destroy()
            return

        messagebox.showinfo("Information", "This will take a while. Press OK and then wait for another popup.")
        self.parent_window.lift()
        progress_window.lift()
        
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        subfolder_name = "pairwise overlaps"
        folder = os.path.join(out_dir, subfolder_name)
        os.makedirs(folder, exist_ok=True)
        csv_file = os.path.join(folder, f"pairwise_overlaps_{timestamp}.csv")
        gexf_file = os.path.join(folder, f"tag_network_{timestamp}.gexf")

        # ----------------------------
        # Step 2: Compute pairwise overlaps
        # ----------------------------
        progress_label.config(text="Computing overlaps...")
        progress_bar['value'] = 10
        progress_window.update_idletasks()

        overlaps = np.zeros((n_tags, n_tags), dtype=np.int32)

        for i in range(n_tags):
            set_i = tag_to_verses[all_tags[i]]
            for j in range(i + 1, n_tags):
                set_j = tag_to_verses[all_tags[j]]
                count = len(set_i & set_j)
                overlaps[i, j] = overlaps[j, i] = count
            if (i + 1) % 10 == 0:
                progress_bar['value'] = 10 + 80 * (i / n_tags)
                progress_label.config(text=f"Processed {i+1}/{n_tags} tags...")
                progress_window.update_idletasks()

        # ----------------------------
        # Step 3: Export CSV
        # ----------------------------
        progress_label.config(text="Writing CSV file...")
        progress_bar['value'] = 95
        progress_window.update_idletasks()

        with open(csv_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Tag1", "Tag2", "OverlapCount"])
            for i in range(n_tags):
                for j in range(i + 1, n_tags):
                    c = overlaps[i, j]
                    if c > 0:
                        writer.writerow([all_tags[i], all_tags[j], c])

        # ----------------------------
        # Step 4: Export network file (GEXF)
        # ----------------------------
        progress_label.config(text="Building network graph...")
        progress_window.update_idletasks()

        G = nx.Graph()
        for t in all_tags:
            G.add_node(t)
        for i in range(n_tags):
            for j in range(i + 1, n_tags):
                c = overlaps[i, j]
                if c > 0:
                    # Edge weight = number of overlapping verses
                    G.add_edge(all_tags[i], all_tags[j], weight=int(c))

        nx.write_gexf(G, gexf_file)

        # ----------------------------
        # Done
        # ----------------------------
        progress_bar['value'] = 100
        progress_label.config(text="Done.")
        progress_window.update_idletasks()
        time.sleep(0.3)
        progress_window.destroy()

        messagebox.showinfo("All done!",
                            f"Pairwise overlaps saved to:\n{csv_file}\n\n"
                            f"Network file saved to:\n{gexf_file}\n\n"
                            "You can open the .gexf file in Gephi for interactive exploration.")
        self.parent_window.lift()

    def export_tag_verse_matrix(self, tags=None, verses=None):
        """
        Export a spreadsheet with synonym-group columns and verse rows
        marking which verses belong to which tag-groups.
        """

        # Ask user for base folder
        xlsx_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx"), ("All files", "*.*")])
        #out_dir = filedialog.askdirectory(title="Select folder for Tag–Verse matrix export")
        if not xlsx_file:
            return
        self.parent_window.lift()

        #import datetime
        #timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        #subfolder_name = f"tag_verse_matrix_{timestamp}"
        #folder = os.path.join(out_dir, subfolder_name)
        #os.makedirs(folder, exist_ok=True)

        # -----------------
        # 1. Build verse_refs from input tags/verses
        # -----------------
        verse_refs = set()

        if verses is not None and len(verses) > 0:
            for vtuple in verses:
                vref = bdblib.normalize_vref({
                    "ID": None,
                    "start_book": vtuple[0],
                    "start_chapter": vtuple[1],
                    "start_verse": vtuple[2],
                    "end_book": vtuple[3],
                    "end_chapter": vtuple[4],
                    "end_verse": vtuple[5]
                })
                verse_refs.add(vref)
        elif tags is not None and len(tags) > 0:
            for t in tags:
                stack = [t]
                seen_tags = set()
                while stack:
                    tg = stack.pop()
                    if tg in seen_tags:
                        continue
                    seen_tags.add(tg)
                    for row in bdblib.get_db_stuff(self.dbdata, "verse", "tag", tg):
                        vref = bdblib.normalize_vref(row)
                        verse_refs.add(vref)
                    syns = [b['tag'] for b in bdblib.get_db_stuff(self.dbdata, "tag", "tag", tg)]
                    for s in syns:
                        if s not in seen_tags:
                            stack.append(s)
        else:
            messagebox.showinfo("No Data", "No verses or tags to analyze.")
            self.parent_window.lift()
            return

        if not verse_refs:
            messagebox.showinfo("No Data", "No verses found for these tags.")
            self.parent_window.lift()
            return

        # -----------------
        # 2. Build verse_to_tags + collect all tags present
        # -----------------
        all_tags_in_verses = set()
        verse_to_tags = defaultdict(set)
        for vref in verse_refs:
            rows = bdblib.get_db_stuff(self.dbdata, "tag", "verse", vref)
            for r in rows:
                t = r['tag']
                verse_to_tags[vref].add(t)
                all_tags_in_verses.add(t)

        # -----------------
        # 3. Build synonym groups
        # -----------------
        visited = set()
        syn_groups = []
        for t in all_tags_in_verses:
            if t in visited:
                continue
            stack = [t]
            comp = set()
            while stack:
                tg = stack.pop()
                if tg in visited:
                    continue
                visited.add(tg)
                comp.add(tg)
                syns = [b['tag'] for b in bdblib.get_db_stuff(self.dbdata, "tag", "tag", tg)]
                for s in syns:
                    if s in all_tags_in_verses and s not in visited:
                        stack.append(s)
            syn_groups.append(comp)

        if not syn_groups:
            messagebox.showinfo("No Data", "No tag groups found in these verses.")
            return

        # For each synonym group, find verses where ANY tag in the group occurs
        group_to_verses = []
        for g in syn_groups:
            vs = {vref for vref, tagsset in verse_to_tags.items() if tagsset & g}
            group_to_verses.append((g, vs))

        # Sort groups by count of verses
        group_to_verses.sort(key=lambda x: len(x[1]), reverse=True)

        # Build display labels
        group_labels = []
        for g, vs in group_to_verses:
            first_tag = sorted(g)[0]
            label = f"{first_tag} (+{len(g)-1} syns)" if len(g) > 1 else first_tag
            group_labels.append(label)

        # -----------------
        # 4. Build the spreadsheet
        # -----------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Tag–Verse Matrix"

        # Header row: Verse + each tag-group
        ws.cell(row=1, column=1, value="Verse")
        for col_idx, label in enumerate(group_labels, start=2):
            ws.cell(row=1, column=col_idx, value=label)

        # Data rows
        verse_list_sorted = sorted(list(verse_refs))
        for row_idx, vref in enumerate(verse_list_sorted, start=2):
            ws.cell(row=row_idx, column=1, value=vref)
            for col_idx, (group, vs) in enumerate(group_to_verses, start=2):
                val = 1 if vref in vs else 0
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Save
        #xlsx_file = os.path.join(folder, f"tag_verse_matrix_{timestamp}.xlsx")
        wb.save(xlsx_file)

        messagebox.showinfo(
            "Export Complete",
            f"Tag–Verse matrix exported to:\n{xlsx_file}"
        )
        self.parent_window.lift()



    def display_attributes(self, dbstuff = None, canvas_width = None):
        if canvas_width == None:
            canvas_width = self.canvas_width
        else:
            self.canvas_width = canvas_width
        y_offset = 40
        x_offset = 5
        textelbowroom = 10
        textlinegap = 2
        textlineheight = Font.metrics(self.canvasFont)["linespace"]
        boldlineheight = Font.metrics(self.boldFont)["linespace"]
        italiclineheight = Font.metrics(self.italicFont)["linespace"]
        right_x_offset = 30
        panelWidth = canvas_width - right_x_offset

        #Clear the canvas
        self.canvas.delete("all")

        #show a header?
        

        if dbstuff != None:
            self.dbdata = dbstuff
        else:
            dbstuff = self.dbdata

        #if there's no db file, then don't show anything.
        if self.dbdata is None:
            self.canvas.create_text(x_offset, y_offset, text="Go select a verse and load a db first.", fill="green", font = self.canvasFont)
        else:

            # PREP LIST OF TAGS AND VERSES ##---- DONE
            #print("test 1")
            #delete tag button size...
            
            checklist = self.tags_list.copy() #in this case, the checklist is all strings, so no tag['tag'] or tag['id']
            synonymlist = [] #the list of tags that will actually be shown
            verses = []  # To hold the final set of verses

            if len(self.tags_list) > 0:
                grouped_verses = []  # To hold verse sets for each tag + synonyms group

                # Go through each selected tag in self.tags_list
                for tag in self.tags_list:
                    #prepare the verse set for this tag and its synonyms by getting the verses for this tag first.
                    verse_set = set()
                    #print("tag", tag)
                    vd = bdblib.get_db_stuff(self.dbdata, "verse", "tag", tag) #vd now has a list of all the verses as returned by get_db_stuff
                    vdcopy = vd.copy() #make a copy of verse data so we can parse it safely
                    #eliminate all verses that aren't in the selected books
                    if (len(self.selected_books) > 0):
                        for verse in vdcopy:
                            if (bdblib.book_proper_names[verse["start_book"]] not in bdblib.book_proper_names) and (bdblib.book_proper_names[verse["end_book"]] not in bdblib.book_proper_names):
                                vd.remove(verse)
                        vdcopy = vd.copy() #remake vdcopy for the next filter


                    #print("vd 1", vd)
                    #convert the verse data to a tuple of numbers to make it hashable for some list functions
                    verse_set.update(((m["start_book"],m["start_chapter"],m["start_verse"],m["end_book"],m["end_chapter"],m["end_verse"]) for m in vd))

                    # Get all synonyms for the current tag (including the tag itself)
                    synonyms = bdblib.get_db_stuff(self.dbdata, "tag", "tag", tag)
                    #for synonym in synonyms:
                    while len(synonyms) > 0:
                        synonym = synonyms.pop()
                        #make the synonymlist which will be used for the tags
                        if synonym['tag'] not in checklist:
                            #synonyms won't have a delete button on the display, so we group them in the list in order to make it clear what they are.
                            index = checklist.index(tag)
                            checklist.insert(index+1,synonym['tag'])
                            synonymlist.append(synonym['tag'])
                            synonyms += bdblib.get_db_stuff(self.dbdata, "tag", "tag", synonym['tag'])
                        # add the verses for every synonym
                        vd = bdblib.get_db_stuff(self.dbdata, "verse", "tag", synonym['tag'])
                        #print("vd 2", vd)
                        verse_set.update(((m["start_book"],m["start_chapter"],m["start_verse"],m["end_book"],m["end_chapter"],m["end_verse"]) for m in vd))

                    # Append the verse set (tag + its synonyms) to grouped_verses
                    grouped_verses.append(verse_set)

                # This next section is why I wanted the hashable tuples....
                # Union logic: collect all unique verses from all tag groups
                if self.union:
                    verses = set()
                    # making it a set will prevent duplicate entries
                    for verse_set in grouped_verses:
                        verses.update(verse_set)
                # Intersection logic: collect only verses present in all tag groups
                elif self.intersection:
                    verses = grouped_verses[0]  # Start with the first group of verses (this is a set, not a list)
                    for verse_set in grouped_verses[1:]:
                        verses = verses.intersection(verse_set)  # Keep only common verses
                # symmetric difference logic: show only verses which are present in only one of the tag groups
                elif self.symmetric_difference:
                    verses = grouped_verses[0]
                    for verse_set in grouped_verses[1:]:
                        verses = verses.symmetric_difference(verse_set)
                
                #both union and intersection converted verses to a set. Make it a list again.
                verses = list(verses)
            
            #some statistical data...
            showtext = "total verse count: " +str(len(self.verse_xref_list))
            versecount_text = self.canvas.create_text(x_offset+textelbowroom, y_offset+textelbowroom, text=showtext, anchor=tk.NW, font=self.canvasFont)
            y_offset += textlineheight + textlinegap*2 + textelbowroom*2
            
            # HEADER ##---- DONE
            if self.union:
                title_text = "=== UNION: Showing all verses for selected tags ==="
            elif self.intersection:
                title_text = "=== INTERSECTION: Showing verses shared by all selected tags ==="
            elif self.symmetric_difference:
                title_text = "=== SYMMETRIC DIFFERENCE: Showing verses unique to only one of the selected tags ==="
            self.canvas.create_text(x_offset, y_offset, text=title_text, anchor=tk.W, font=self.boldFont)
            y_offset += boldlineheight + textlinegap
            
            buttonX = x_offset
            
            # ====== Export Dropdown ======
            # Use a tkinter OptionMenu or ttk.Combobox for export options

            export_options = [
                "Export All Tag/Note/Verses",
                "Export All Verse Notes",
                "Export Single Topic Files",
                "Export Tag Networks",
                "Export Subtopic Breakdowns",
                "Export Pairwise Overlaps",
                "Export Tag Verse Matrix"
            ]
            self.selected_export_option = StringVar()
            self.selected_export_option.set(export_options[0])  # default

            # We create an actual widget over the canvas for dropdown (simplest)
            self.export_dropdown = tk.OptionMenu(self.master, self.selected_export_option, *export_options)
            self.canvas.create_window(buttonX, y_offset, anchor=tk.NW, window=self.export_dropdown)
            buttonX += 200  # width of dropdown

            # "Go" button
            def run_selected_export():
                opt = self.selected_export_option.get()
                if opt == "Export All Tag/Note/Verses":
                    self.export_tags_and_synonyms(None)
                elif opt == "Export All Verse Notes":
                    self.export_verse_notes(None)
                elif opt == "Export Single Topic Files":
                    self.export_single_topics(None)
                elif opt == "Export Tag Networks":
                    self.export_tag_networks(None)
                elif opt == "Export Subtopic Breakdowns":
                    self.export_subtopic_breakdowns(None)
                elif opt == "Export Pairwise Overlaps":
                    self.export_pairwise_overlaps(None)
                elif opt == "Export Tag Verse Matrix":
                    self.export_tag_verse_matrix(tags=self.tags_list, verses=verses)

            go_button = tk.Button(self.master, text="Go", command=run_selected_export)
            self.canvas.create_window(buttonX, y_offset, anchor=tk.NW, window=go_button)
            buttonX += 50
            
            #new row
            y_offset += textlineheight + 2*textelbowroom
            buttonX = x_offset

        # ====== MAIN BUTTON ROW (toggle / copy / analyze) ======
            
            # Toggle Set Operation
            buttonText = "Toggle Set Operation"
            button_width = self.canvasFont.measure(buttonText) + 2*textelbowroom
            self.create_toggle_union_btn = self.canvas.create_rectangle(buttonX, y_offset,
                buttonX + button_width, y_offset + textlineheight + 2*textelbowroom,
                fill='snow', tags='toggle_union_button')
            self.canvas.create_text(buttonX+textelbowroom, y_offset+textelbowroom,
                text=buttonText, anchor=tk.NW, font=self.canvasFont, tags='toggle_union_button')
            self.canvas.tag_bind('toggle_union_button', '<Button-1>', lambda event: self.toggle_union(event))
            buttonX += x_offset + button_width

            # Copy These Verses
            buttonText = "Copy These Verses"
            button_width = self.canvasFont.measure(buttonText) + 2*textelbowroom
            self.copy_this_verselist_button = self.canvas.create_rectangle(buttonX, y_offset,
                buttonX + button_width, y_offset + textlineheight + 2*textelbowroom,
                fill='snow', tags='copy_vlist_button')
            self.canvas.create_text(buttonX+textelbowroom, y_offset+textelbowroom,
                text=buttonText, anchor=tk.NW, font=self.canvasFont, tags='copy_vlist_button')
            self.canvas.tag_bind('copy_vlist_button', '<Button-1>', lambda event: self.copy_verse_lists(event))
            buttonX += x_offset + button_width

            #new row
            y_offset += 10 + textlineheight + 2*textelbowroom
            buttonX = x_offset
            
        #Tag Sorting Buttons:
            buttonText = "Search Tag"
            button_width = self.canvasFont.measure(buttonText) + 2*textelbowroom
            self.create_tag_button_rect = self.canvas.create_rectangle(buttonX, y_offset, buttonX + button_width, y_offset + textlineheight + 2*textelbowroom, fill='azure', tags='create_tag_button')
            self.canvas.create_text(buttonX+textelbowroom, y_offset+textelbowroom, text=buttonText, anchor=tk.NW, font=self.canvasFont, tags = 'create_tag_button')#button text for open db
            self.canvas.tag_bind('create_tag_button', '<Button-1>', lambda event: self.select_tag(event))
            buttonX += x_offset + button_width
            
            buttonText = "Select Books"
            button_width = self.canvasFont.measure(buttonText) + 2*textelbowroom
            self.select_book_rect = self.canvas.create_rectangle(buttonX, y_offset, buttonX + button_width, y_offset + textlineheight + 2*textelbowroom, fill='azure', tags='select_book_button')
            self.canvas.create_text(buttonX+textelbowroom, y_offset+textelbowroom, text=buttonText, anchor=tk.NW, font=self.canvasFont, tags = 'select_book_button')#button text for open db
            self.canvas.tag_bind('select_book_button', '<Button-1>', lambda event: self.select_book(event))
            buttonX += x_offset + button_width

            # Analyze Subtopic Tags (new)
            buttonText = "Analyze Subtopics"
            button_width = self.canvasFont.measure(buttonText) + 2*textelbowroom
            self.analyze_subtopic_button = self.canvas.create_rectangle(buttonX, y_offset,
                buttonX + button_width, y_offset + textlineheight + 2*textelbowroom,
                fill='snow', tags='analyze_subtopic_button')
            self.canvas.create_text(buttonX+textelbowroom, y_offset+textelbowroom,
                text=buttonText, anchor=tk.NW, font=self.canvasFont, tags='analyze_subtopic_button')
            if self.intersection:
                #Let's go ahead and bind the analysis button with our verse listing
                self.canvas.tag_bind(
                    'analyze_subtopic_button', 
                    '<Button-1>',
                    lambda event: self.analyze_subtopic_tags(event, tags=self.tags_list, verses=verses)
                )
            else:
                def sortmode_complain():
                    messagebox.showinfo("Sort Mode", "This feature only works with set operation: intersection")
                    self.parent_window.lift()
                self.canvas.tag_bind(
                    'analyze_subtopic_button', 
                    '<Button-1>',
                    lambda event: sortmode_complain()
                )
            buttonX += x_offset + button_width
            
            #new row
            y_offset += 10 + textlineheight + 2*textelbowroom
            buttonX = x_offset

            # comments only checkbox -- the filtering feature associated with this is not written yet
            if self.comments_only:
                comments_checkbox_fill = "black"
            else:
                comments_checkbox_fill = "white"

            self.canvas.create_rectangle(buttonX, y_offset, buttonX+textlineheight, y_offset+textlineheight, outline="black", fill=comments_checkbox_fill)
            self.checkbox_id = self.canvas.create_rectangle(buttonX+textlineheight/4, y_offset+textlineheight/4, buttonX+(textlineheight*3/4), y_offset+(textlineheight*3/4), fill=comments_checkbox_fill)
            self.label_id = self.canvas.create_text(buttonX+textlineheight+textelbowroom, y_offset, text="Only show verses with comments", anchor=tk.NW)
            self.canvas.tag_bind(self.checkbox_id, "<Button-1>", self.toggle_comments_only)
            self.canvas.tag_bind(self.label_id, "<Button-1>", self.toggle_comments_only)

            #new row...
            y_offset += 10 + textlineheight + 2*textelbowroom
            buttonX = x_offset
                

            #SHOW THE BOOK FILTER LIST
            xb_width = self.canvasFont.measure(" x ")
            tagx = 0
            
            for book in self.selected_books:
                #print(book)
                book_width = self.canvasFont.measure(book) + 2*textelbowroom
                if book_width + xb_width + tagx+1 + x_offset + right_x_offset > panelWidth: #+1 for that thicker line between tags
                    tagx = 0
                    y_offset += textlineheight + textlinegap*3
                #every book gets an x
                book_delete_binder = "Delete_" + str(book)
                tagx += 1 #making a thicker line between tag groups
                self.canvas.create_rectangle(x_offset+tagx, y_offset, x_offset+tagx+xb_width, y_offset+textlineheight+textlinegap*2, fill='coral1', tags=book_delete_binder)
                self.canvas.create_text(x_offset+tagx, y_offset+textlinegap, text=" X", anchor=tk.NW, font=self.canvasFont, tags=book_delete_binder)
                
                tagx += xb_width
                #clicking the book deletes the book from the list
                self.canvas.create_rectangle(x_offset+tagx, y_offset, x_offset+tagx+book_width, y_offset+textlineheight+textlinegap*2, fill='azure', tags=book_delete_binder)
                self.canvas.create_text(x_offset+tagx, y_offset+textlinegap, text=" " + book, anchor=tk.NW, font=self.canvasFont, tags=book_delete_binder)
                
                self.canvas.tag_bind(book_delete_binder, '<Button-1>', lambda event, mytag=book: self.delete_book(event, book))
                tagx += book_width


            if len(self.selected_books) > 0:
                y_offset += textlineheight + textlinegap*3 + 10
            
            #SHOW THE TAGS LIST
            tagx = 0
            for tag in checklist:
                tag_width =  self.canvasFont.measure(tag) + 2*textelbowroom

                #wrap tag list
                if tag_width + xb_width + tagx+1 + x_offset + right_x_offset > panelWidth: #+1 for that thicker line between tags
                    tagx = 0
                    y_offset += textlineheight + textlinegap*3

                #tag delete "X" button.
                #only do this if this is not a synonym
                if tag not in synonymlist:
                    tag_delete_binder = "Delete_" + str(checklist.index(tag))
                    tagx += 1 #making a thicker line between tag groups
                    self.canvas.create_rectangle(x_offset+tagx, y_offset, x_offset+tagx+xb_width, y_offset+textlineheight+textlinegap*2, fill='coral1', tags=tag_delete_binder)
                    self.canvas.create_text(x_offset+tagx, y_offset+textlinegap, text=" X", anchor=tk.NW, font=self.canvasFont, tags=tag_delete_binder)
                    self.canvas.tag_bind(tag_delete_binder, '<Button-1>', lambda event, mytag=tag: self.delete_tag(event, mytag))
                    tagx += xb_width
                #actual tag display
                tag_display_binder = "Display_" + str(checklist.index(tag))
                self.canvas.create_rectangle(x_offset+tagx, y_offset, x_offset+tagx+tag_width, y_offset+textlineheight+textlinegap*2, fill='azure', tags=tag_display_binder)

                #if there's a note on this tag, put a little triangle on the corner of the tag's button
                if (len(bdblib.get_db_stuff(self.dbdata, "note", "tag", tag)) > 0):
                    self.canvas.create_polygon( #triangle for tags with comments
                        x_offset + tagx + tag_width,  # x1: Upper-right corner x
                        y_offset,                     # y1: Upper-right corner y
                        x_offset + tagx + tag_width - 10,  # x2: 10 pixels left of upper-right corner x
                        y_offset,                     # y2: Same y as upper-right corner
                        x_offset + tagx + tag_width,  # x3: Upper-right corner x
                        y_offset + 10,                 # y3: 10 pixels down from upper-right corner y
                        fill='black',
                        outline='black',  # Optional: remove or change if you don't want an outline
                        tags=tag_display_binder
                    )
                self.canvas.create_text(x_offset+tagx, y_offset+textlinegap, text=" " + tag, anchor=tk.NW, font=self.canvasFont, tags=tag_display_binder)
                #button event to show tag details in this panel...
                self.canvas.tag_bind(tag_display_binder, '<Button-1>', lambda event, item_id = tag : self.on_tag_click(event,item_id))
                tagx += tag_width

            y_offset += textlineheight + textlinegap*3 + 10

            # SHOW THE VERSE LIST
            
            try:
                # If the user makes a DB using a version of the Bible that has the apocrypha, and then tries to pull notes about Revelation from that DB while he has a protestant Bible loaded...
                #    then the reference to bdblib_Lib.book_proper_names[] will throw an index out of range error.
                #    I am making this tool primarily for myself to use, and I don't include the apocrypha in my Bible, so I don't plan to fix this.
                
                verses.sort(key=lambda r: (r[0], r[1], r[2]))#sort by start book first, then chapter, then verse, so all the verses appear in order
                self.verse_xref_list = [(bdblib.book_proper_names[q[0]]+" "+str(q[1])+":"+str(q[2]), bdblib.book_proper_names[q[3]]+" "+str(q[4])+":"+str(q[5])) for q in verses]
            except:
                self.verse_xref_list = []
                print("Failed to get the verse references for that tag. This error might occur if your DB was made with a version of the Bible that had different books from the version you're currently using. For example, if the DB was made including the apocrypha, but your current Bible doesn't have it.")
            self.canvas.itemconfigure(versecount_text, text = "total verse count: " +str(len(self.verse_xref_list)))
            tagx = 0
            self.shown_verses = []
            for xverse in self.verse_xref_list:
                itemText = combineVRefs(xverse[0],xverse[1])#this is the human readable verse reference
                #Check if we're filtering on verses with comments
                if self.comments_only and not (len(bdblib.get_db_stuff(self.dbdata, "note", "verse", itemText)) > 0):
                    continue #if we're filtering and no comment, then skip this iteration in the foor loop.
                #check if the verse should be book filtered
                if len(self.selected_books) > 0:
                    skipverse = True
                    for book in self.selected_books:
                        if book in itemText:
                            skipverse = False
                            break
                    if skipverse:
                        continue

                self.shown_verses.append(itemText + ", ")
                #draw the button and verse
                button_width = self.canvasFont.measure(itemText) + 2*textelbowroom
                if button_width + tagx + x_offset + right_x_offset > panelWidth:
                    tagx = 0
                    y_offset += textlineheight + textlinegap*3
                clicktag = "verse_click_"+str(tagx)+"_"+str(y_offset)
                self.canvas.create_rectangle(x_offset+tagx, y_offset, x_offset+tagx+button_width, y_offset+textlineheight+textlinegap*2, fill='azure', tags=clicktag)
                self.canvas.create_text(x_offset+tagx+textelbowroom, y_offset+textlinegap, text=itemText, anchor=tk.NW, font = self.canvasFont, tags=clicktag)
                self.canvas.tag_bind(clicktag, '<Button-1>', lambda event, payload=xverse: self.on_verse_click(event, payload))
                tagx += button_width
            #print("test 3")
            y_offset += textlineheight + textlinegap*3 + 10
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))


